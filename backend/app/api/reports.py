from fastapi import APIRouter, Depends, HTTPException, Header
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
from io import BytesIO
from ..database import get_db
from ..models.grade import Grade
from ..models.student import Student
from ..models.user import User
from ..models.group import Group
from ..models.subject import Subject
from ..models.teacher import Teacher
from ..models.attendance import Attendance
from ..utils.security import verify_token
from ..services.export_service import export_grades_to_excel

router = APIRouter()


def get_current_user(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    token = authorization.replace("Bearer ", "")
    payload = verify_token(token)
    if not payload:
        raise HTTPException(status_code=401, detail="Недействительный токен")
    return int(payload["sub"]), payload["role"]


@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    return {
        "students": db.query(Student).count(),
        "teachers": db.query(Teacher).count(),
        "groups": db.query(Group).count(),
        "subjects": db.query(Subject).count(),
        "grades": db.query(Grade).count(),
        "attendance": db.query(Attendance).count()
    }


@router.get("/group-report")
def get_group_report(
        group_id: int,
        subject_id: int,
        date_from: str = None,
        date_to: str = None,
        db: Session = Depends(get_db)
):
    group = db.query(Group).filter(Group.id == group_id).first()
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not group or not subject:
        raise HTTPException(status_code=404, detail="Группа или предмет не найдены")

    students = db.query(Student).filter(Student.group_id == group_id).all()
    result = []
    for student in students:
        user = db.query(User).filter(User.id == student.user_id).first()
        query = db.query(Grade).filter(Grade.student_id == student.id, Grade.subject_id == subject_id)
        if date_from: query = query.filter(Grade.date >= date_from)
        if date_to: query = query.filter(Grade.date <= date_to)
        grades = query.order_by(Grade.date.desc()).all()
        avg = sum(g.value for g in grades) / len(grades) if grades else 0
        result.append({
            "student_id": student.id,
            "student_name": user.full_name if user else "-",
            "student_card": student.student_card_number,
            "grades": [{"id": g.id, "value": g.value, "type": g.type, "date": g.date.isoformat() if g.date else None, "comment": g.comment} for g in grades],
            "average": round(avg, 2),
            "total_grades": len(grades)
        })
    return {"group_name": group.name, "subject_name": subject.name, "students": result}


@router.get("/group-report/excel")
def export_group_report(group_id: int, subject_id: int, date_from: str = None, date_to: str = None, db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not group or not subject:
        raise HTTPException(status_code=404, detail="Группа или предмет не найдены")

    students = db.query(Student).filter(Student.group_id == group_id).all()
    all_grades = []
    for student in students:
        user = db.query(User).filter(User.id == student.user_id).first()
        query = db.query(Grade).filter(Grade.student_id == student.id, Grade.subject_id == subject_id)
        if date_from: query = query.filter(Grade.date >= date_from)
        if date_to: query = query.filter(Grade.date <= date_to)
        for g in query.all():
            all_grades.append({"value": g.value, "type": g.type, "date": g.date, "comment": g.comment, "student_name": user.full_name if user else "-", "subject_name": subject.name})

    excel_file = export_grades_to_excel(all_grades, f"{group.name} - {subject.name}")
    return StreamingResponse(excel_file, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=vedomost.xlsx', 'Access-Control-Allow-Origin': '*'})


@router.get("/audit-logs")
def get_audit_logs(username: str = None, action: str = None, date_from: str = None, date_to: str = None, page: int = 1, page_size: int = 20, db: Session = Depends(get_db)):
    from ..models.audit_log import AuditLog
    query = db.query(AuditLog)
    if username: query = query.filter(AuditLog.username.ilike(f"%{username}%"))
    if action: query = query.filter(AuditLog.action.ilike(f"%{action}%"))
    if date_from: query = query.filter(AuditLog.created_at >= date_from)
    if date_to: query = query.filter(AuditLog.created_at <= date_to)
    total = query.count()
    logs = query.order_by(AuditLog.created_at.desc()).offset((page-1)*page_size).limit(page_size).all()
    return {
        "total": total, "page": page, "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "data": [{"id": l.id, "user_id": l.user_id, "username": l.username, "action": l.action, "details": l.details, "created_at": l.created_at.isoformat() if l.created_at else None} for l in logs]
    }


@router.get("/audit-logs/excel")
def export_audit_logs(username: str = None, action: str = None, date_from: str = None, date_to: str = None, db: Session = Depends(get_db)):
    from ..models.audit_log import AuditLog
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    query = db.query(AuditLog)
    if username: query = query.filter(AuditLog.username.ilike(f"%{username}%"))
    if action: query = query.filter(AuditLog.action.ilike(f"%{action}%"))
    if date_from: query = query.filter(AuditLog.created_at >= date_from)
    if date_to: query = query.filter(AuditLog.created_at <= date_to)

    logs = query.order_by(AuditLog.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Логи"

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='6c5ce7', end_color='6c5ce7', fill_type='solid')

    for col, h in enumerate(['ID', 'Пользователь', 'Действие', 'Детали', 'Дата'], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, l in enumerate(logs, 2):
        ws.cell(row=i, column=1, value=l.id).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=2, value=l.username)
        ws.cell(row=i, column=3, value=l.action)
        ws.cell(row=i, column=4, value=l.details or '-')
        ws.cell(row=i, column=5, value=l.created_at.strftime('%d.%m.%Y %H:%M') if l.created_at else '-')

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=logs.xlsx'})