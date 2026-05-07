from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def export_grades_to_excel(grades_data: list, student_name: str = "Студент") -> BytesIO:
    """
    Экспорт оценок в Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Оценки"

    # Стили
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4361ee', end_color='4361ee', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовок
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Ведомость оценок — {student_name}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2c3e50')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Дата выгрузки: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=10, color='7f8c8d')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Шапка таблицы
    headers = ['№', 'Предмет', 'Оценка', 'Тип', 'Дата', 'Комментарий']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Данные
    type_map = {'exam': 'Экзамен', 'test': 'Зачёт', 'coursework': 'Курсовая', 'practice': 'Практика'}

    for i, grade in enumerate(grades_data, 1):
        row = i + 4

        # Поддержка и объектов, и словарей
        if isinstance(grade, dict):
            subject_name = grade.get('subject_name', f"ID: {grade.get('subject_id', '-')}")
            value = grade.get('value', '-')
            grade_type = grade.get('type', '-')
            date_val = grade.get('date', '-')
            comment = grade.get('comment', '-') or '-'
        else:
            subject_name = getattr(grade, 'subject_name', None) or f"ID: {getattr(grade, 'subject_id', '-')}"
            value = getattr(grade, 'value', '-')
            grade_type = getattr(grade, 'type', '-')
            date_val = getattr(grade, 'date', '-')
            comment = getattr(grade, 'comment', None) or '-'

        if hasattr(date_val, 'strftime'):
            date_val = date_val.strftime('%d.%m.%Y')
        elif isinstance(date_val, str) and len(date_val) >= 10:
            date_val = date_val[:10]

        type_map = {'exam': 'Экзамен', 'test': 'Зачёт', 'coursework': 'Курсовая', 'practice': 'Практика'}
        grade_type_ru = type_map.get(grade_type, grade_type)

        value_fill = PatternFill()
        if isinstance(value, (int, float)):
            if value >= 5:
                value_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
            elif value >= 4:
                value_fill = PatternFill(start_color='d1ecf1', end_color='d1ecf1', fill_type='solid')
            elif value >= 3:
                value_fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
            else:
                value_fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')

        row_data = [i, subject_name, value, grade_type_ru, date_val, comment]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if col == 3:
                cell.fill = value_fill
                cell.font = Font(bold=True)

    # Ширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30

    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_attendance_to_excel(attendance_data: list, student_name: str = "Студент") -> BytesIO:
    """
    Экспорт посещаемости в Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Посещаемость"

    # Стили
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='4361ee', end_color='4361ee', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовок
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Ведомость посещаемости — {student_name}"
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2c3e50')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = f"Дата выгрузки: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=10, color='7f8c8d')
    ws['A2'].alignment = Alignment(horizontal='center')

    # Статистика
    status_map = {'present': 'Присутствовал', 'absent': 'Отсутствовал',
                  'late': 'Опоздал', 'excused': 'Уваж. причина'}

    total = len(attendance_data)
    present = sum(1 for r in attendance_data if getattr(r, 'status', None) == 'present')
    absent = sum(1 for r in attendance_data if getattr(r, 'status', None) == 'absent')
    rate = round(present / total * 100, 1) if total > 0 else 0

    ws.merge_cells('A4:E4')
    ws['A4'] = f"Всего занятий: {total} | Посещено: {present} | Пропущено: {absent} | Процент: {rate}%"
    ws['A4'].font = Font(name='Arial', bold=True, size=11)

    # Шапка таблицы
    headers = ['№', 'Предмет', 'Дата', 'Статус', 'Отметка']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for i, record in enumerate(attendance_data, 1):
        row = i + 6

        subject_id = getattr(record, 'subject_id', '-')
        date_val = getattr(record, 'date', '-')
        status = getattr(record, 'status', '-')

        if hasattr(date_val, 'strftime'):
            date_val = date_val.strftime('%d.%m.%Y')
        elif isinstance(date_val, str) and len(date_val) >= 10:
            try:
                date_val = date_val[:10]
            except:
                pass

        status_text = status_map.get(status, status)

        # Цвет статуса
        status_fill = PatternFill()
        if status == 'present':
            status_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
        elif status == 'absent':
            status_fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
        elif status == 'late':
            status_fill = PatternFill(start_color='fff3cd', end_color='fff3cd', fill_type='solid')
        elif status == 'excused':
            status_fill = PatternFill(start_color='d1ecf1', end_color='d1ecf1', fill_type='solid')

        row_data = [i, f"Предмет ID: {subject_id}", date_val, status_text, '']
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = cell_alignment
            cell.border = thin_border
            if col == 4:
                cell.fill = status_fill

    # Ширина
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output